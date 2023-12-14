#excel failiem

from openpyxl import workbook, load_workbook
saraksts=[]
wb=load_workbook('nosaukums.xclc')
Ws=Wb.active

Max_row=ws.Max_row
For row in range(2, max_row+1):
r_data=[]
C_value=ws['C'+str(row)].value
E_value=int(ws['E'+str(row)].value)
r_data.append(C_value)
R_data.append(E_value)
Saraksts.append(r_data)
Wb.close()

for row in Saraksts:
    print(row[0])
    print(row[1])
    
Jauns=[]
for row in saraksts:
    if(row=="IT"):
    Jauns.append(row[1])
    print(row[1])